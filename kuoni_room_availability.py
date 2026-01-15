#!/usr/bin/env python3
"""Fetch Kuoni HotelMap room availability for a list of hotels and emit JSONL,
CSV, Excel, and PNG outputs. Can also convert an existing bookings JSONL into
the tabular/visual formats via --convert-only."""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
from pathlib import Path
from typing import Any, Iterable, Mapping, Sequence

import numpy as np
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image


DEFAULT_URL = os.getenv(
    "KUONI_BOOKINGS_URL",
    "https://api.kuonitumlare.com/v1/hotelmap/bookings/search",
)
DEFAULT_LANGUAGE = "en_GB"
DEFAULT_EVENT_ID = os.getenv("KUONI_BOOKINGS_EVENT_ID", "20000")
DEFAULT_MCODE = os.getenv("KUONI_BOOKINGS_MCODE", "MXXXX")
DEFAULT_AUTH_URL = os.getenv("KUONI_AUTH_URL", "https://kuonitumlare.eu.auth0.com/oauth/token")
DEFAULT_AUTH_CLIENT_ID = os.getenv("KUONI_AUTH_CLIENT_ID")
DEFAULT_AUTH_CLIENT_SECRET = os.getenv("KUONI_AUTH_CLIENT_SECRET")
DEFAULT_AUTH_AUDIENCE = os.getenv("KUONI_AUTH_AUDIENCE", "http://52.208.167.3/staging")
DEFAULT_AUTH_GRANT_TYPE = os.getenv("KUONI_AUTH_GRANT_TYPE", "client_credentials")
DEFAULT_INPUT_CSV = Path("sheets/kuoni_hotel_summary_with_kuoni.csv")
DEFAULT_HOTEL_COLUMN = "kuoni_id"
DEFAULT_JSONL = Path("output_sheets/bookings_search_results.jsonl")
DEFAULT_CSV = Path("output_sheets/bookings_search_availability.csv")
DEFAULT_EXCEL = Path("output_sheets/bookings_search_availability.xlsx")
DEFAULT_PNG = Path("output_sheets/bookings_search_availability.png")
DEFAULT_PNG_SIZE = (520, 700)
# Edit these to bake in your stay dates/room counts (DATE:ROOMS)
DEFAULT_STAY_ARGS = [
    # "2026-08-30:10",
    # "2026-08-31:9",
]


class KuoniBookingsAPIError(RuntimeError):
    pass


def parse_stays(stays: Sequence[str]) -> list[dict[str, Any]]:
    """Convert CLI --stay arguments like YYYY-MM-DD:ROOMS into payload dicts."""
    parsed: list[dict[str, Any]] = []
    for stay in stays:
        if ":" not in stay:
            raise ValueError(f"Stay '{stay}' must use DATE:ROOMS format")
        date_part, rooms_part = stay.split(":", 1)
        date = date_part.strip()
        if not date:
            raise ValueError(f"Date missing in '{stay}'")
        try:
            rooms = int(rooms_part)
        except ValueError:
            raise ValueError(f"Rooms must be an integer in '{stay}'") from None
        parsed.append({"date": date, "rooms": rooms})
    return parsed


def read_hotel_ids(csv_path: Path, column: str | None) -> list[str]:
    hotel_ids: list[str] = []
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if not reader.fieldnames:
            raise ValueError("CSV appears to be empty")
        target_column = column or reader.fieldnames[0]
        if target_column not in reader.fieldnames:
            raise ValueError(
                f"Column '{target_column}' not found. "
                f"Available columns: {', '.join(reader.fieldnames)}"
            )
        for row in reader:
            raw_value = row.get(target_column)
            if raw_value is None:
                continue
            hotel_id = str(raw_value).strip()
            if hotel_id:
                hotel_ids.append(hotel_id)
    if not hotel_ids:
        raise ValueError(f"No hotel IDs read from column '{target_column}'")
    return hotel_ids


def call_booking_search(
    *,
    session: requests.Session,
    url: str,
    token: str,
    payload: Mapping[str, Any],
    timeout: float,
) -> Mapping[str, Any]:
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }
    response = session.post(url, json=payload, headers=headers, timeout=timeout)
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise KuoniBookingsAPIError(
            f"{response.status_code} error for {payload.get('hotel_id')}: {response.text}"
        ) from exc
    try:
        return response.json()
    except ValueError as exc:
        raise KuoniBookingsAPIError("Non-JSON response received") from exc


def fetch_auth_token(
    session: requests.Session,
    *,
    url: str,
    client_id: str,
    client_secret: str,
    audience: str,
    grant_type: str = "client_credentials",
    timeout: float = 15.0,
) -> str:
    payload = {
        "client_id": client_id,
        "client_secret": client_secret,
        "audience": audience,
        "grant_type": grant_type,
    }
    response = session.post(url, json=payload, timeout=timeout)
    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        raise KuoniBookingsAPIError(
            f"Auth request failed with {response.status_code}: {response.text}"
        ) from exc
    try:
        data = response.json()
    except ValueError as exc:
        raise KuoniBookingsAPIError("Auth endpoint returned non-JSON") from exc
    token = data.get("access_token")
    if not token:
        raise KuoniBookingsAPIError("Auth response missing access_token")
    return str(token)


def iter_payloads(
    hotel_ids: Iterable[str],
    *,
    mcode: str,
    event_id: str,
    best_only: bool,
    language: str,
    config: list[dict[str, Any]] | None,
) -> Iterable[dict[str, Any]]:
    for hotel_id in hotel_ids:
        payload: dict[str, Any] = {
            "mcode": mcode,
            "event_id": event_id,
            "hotel_id": hotel_id,
            "best_only": best_only,
            "language": language,
        }
        if config:
            payload["config"] = config
        yield payload


def _coerce_int(value: Any) -> int | None:
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def extract_hotel_nightly_counts(response: Mapping[str, Any]) -> dict[str, int]:
    """Return {date -> rooms_available} using max-per-date to avoid double counting."""
    nightly: dict[str, int] = {}
    if not response or not response.get("success"):
        return nightly

    def update(date: str | None, count: int | None) -> None:
        if not date or count is None:
            return
        current = nightly.get(date)
        nightly[date] = count if current is None else max(current, count)

    for hotel in response.get("hotels", []) or []:
        for rate in hotel.get("rates", []) or []:
            rate_nightly = rate.get("nightly")
            if isinstance(rate_nightly, Mapping):
                for date, count in rate_nightly.items():
                    update(str(date), _coerce_int(count))

            for room in rate.get("rooms", []) or []:
                for night in room.get("nightly", []) or []:
                    update(night.get("date"), _coerce_int(night.get("rooms")))

    return nightly


def apply_excel_formatting(path: Path) -> None:
    """Insert padding rows and basic styling to match heatmap expectations."""
    wb = load_workbook(path)
    ws = wb.active

    # Insert spacer rows so headers land on row 5 (matches kuoni_generate_sheet layout)
    ws.insert_rows(1, amount=3)
    ws.cell(row=2, column=2).value = "Params"
    ws.cell(row=2, column=3).value = "Bookings availability (rooms)"

    black_fill = PatternFill(fill_type="solid", fgColor="000000")
    grey_fill = PatternFill(fill_type="solid", fgColor="444444")
    white_font = Font(color="FFFFFF")
    orange_font = Font(color="FFA500")

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for col_idx, cell in enumerate(row, start=1):
            cell.fill = black_fill
            cell.font = orange_font if col_idx <= 2 else white_font

    header_row_idx = 5
    if header_row_idx <= max_row:
        for col_idx, cell in enumerate(ws[header_row_idx], start=1):
            cell.fill = grey_fill
            cell.font = orange_font if col_idx <= 2 else white_font

    wb.save(path)


def build_availability_rows(
    hotel_map: Mapping[str, Mapping[str, int]]
) -> tuple[list[str], list[dict[str, object]]]:
    all_dates: set[str] = set()
    for nightly in hotel_map.values():
        all_dates.update(nightly.keys())

    sorted_dates = sorted(all_dates)
    rows: list[dict[str, object]] = []
    for hotel_id in sorted(hotel_map.keys()):
        nightly = hotel_map[hotel_id]
        # Coverage = % of dates with at least 1 room available
        available_nights = 0
        total_nights = len(sorted_dates)
        for date in sorted_dates:
            count = _coerce_int(nightly.get(date))
            if count and count > 0:
                available_nights += 1
        coverage = round(available_nights / total_nights, 4) if total_nights else 0.0
        row: dict[str, object] = {"hotel_id": hotel_id}
        row["coverage"] = coverage
        row.update({date: nightly.get(date, "") for date in sorted_dates})
        rows.append(row)
    return sorted_dates, rows


def write_csv(rows: list[dict[str, object]], dates: list[str], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["hotel_id", "coverage", *dates])
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {len(rows)} hotels to {path}")


def write_excel(rows: list[dict[str, object]], path: Path) -> None:
    df = pd.DataFrame(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(path, index=False)
    apply_excel_formatting(path)
    print(f"Wrote Excel grid to {path}")


def write_png(rows: list[dict[str, object]], size: tuple[int, int], path: Path) -> None:
    grid_df = pd.DataFrame(rows)
    drop_cols = [col for col in ("hotel_id", "coverage") if col in grid_df.columns]
    if drop_cols:
        grid_df = grid_df.drop(columns=drop_cols)
    num_df = grid_df.apply(pd.to_numeric, errors="coerce").fillna(0)
    if num_df.empty or num_df.shape[1] == 0:
        print("No data to render for PNG; skipping.")
        return

    avail = (num_df > 0).astype(np.uint8).to_numpy()
    colors = np.array([[127, 0, 0], [0, 176, 80]], dtype=np.uint8)  # dark red, green
    rgb = colors[avail]
    img = Image.fromarray(rgb, mode="RGB")
    width, height = size
    img = img.resize((width, height), resample=Image.NEAREST)
    path.parent.mkdir(parents=True, exist_ok=True)
    img.save(path)
    print(f"Wrote PNG heatmap to {path}")


def read_jsonl(path: Path) -> Iterable[dict[str, Any]]:
    with path.open() as handle:
        for line in handle:
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except json.JSONDecodeError:
                continue


def build_hotel_map_from_jsonl(path: Path) -> dict[str, dict[str, int]]:
    hotel_map: dict[str, dict[str, int]] = {}
    for record in read_jsonl(path):
        hotel_id = str(record.get("hotel_id") or "").strip()
        if not hotel_id:
            continue
        nightly = extract_hotel_nightly_counts(record.get("response") or {})
        hotel_map[hotel_id] = nightly
    return hotel_map


def generate_outputs(
    hotel_map: Mapping[str, Mapping[str, int]],
    *,
    csv_path: Path | None,
    excel_path: Path | None,
    png_path: Path | None,
    png_size: tuple[int, int],
    no_csv: bool,
    no_excel: bool,
    no_png: bool,
) -> None:
    if not hotel_map:
        print("No hotel data parsed; nothing to write.")
        return

    dates, rows = build_availability_rows(hotel_map)
    if not no_csv and csv_path:
        write_csv(rows, dates, csv_path)
    if not no_excel and excel_path:
        write_excel(rows, excel_path)
    if not no_png and png_path:
        write_png(rows, png_size, png_path)


def fetch_bookings_availability(
    *,
    hotel_ids: Sequence[str],
    config: list[dict[str, Any]] | None,
    args: argparse.Namespace,
    session: requests.Session,
) -> dict[str, dict[str, int]]:
    args.jsonl_output.parent.mkdir(parents=True, exist_ok=True)
    nightly_map: dict[str, dict[str, int]] = {}

    with args.jsonl_output.open("w", encoding="utf-8") as outfile:
        for index, payload in enumerate(
            iter_payloads(
                hotel_ids,
                mcode=args.mcode,
                event_id=args.event_id,
                best_only=args.best_only,
                language=args.language,
                config=config,
            ),
            start=1,
        ):
            hotel_id = payload["hotel_id"]
            print(f"[{index}/{len(hotel_ids)}] {hotel_id}")
            try:
                response = call_booking_search(
                    session=session,
                    url=args.url,
                    token=args.token,
                    payload=payload,
                    timeout=args.timeout,
                )
                record: dict[str, Any] = {"hotel_id": hotel_id, "response": response}
                nightly_map[hotel_id] = extract_hotel_nightly_counts(response)
            except Exception as exc:
                record = {"hotel_id": hotel_id, "error": str(exc)}
                nightly_map.setdefault(hotel_id, {})
                print(f"    [warn] {exc}", file=sys.stderr)

            outfile.write(json.dumps(record) + "\n")

    print(f"Results written to {args.jsonl_output}")
    return nightly_map


def main(argv: Sequence[str] | None = None) -> int:
    load_dotenv()

    parser = argparse.ArgumentParser(
        description=(
            "Fetch Kuoni HotelMap bookings availability for a list of hotels and "
            "render JSONL/CSV/Excel/PNG outputs."
        )
    )
    parser.add_argument(
        "csv_path",
        nargs="?",
        default=DEFAULT_INPUT_CSV,
        type=Path,
        help=f"CSV file containing hotel IDs (default: {DEFAULT_INPUT_CSV})",
    )
    parser.add_argument(
        "--hotel-column",
        default=DEFAULT_HOTEL_COLUMN,
        help=(
            "Column name in the CSV that holds the Kuoni hotel IDs "
            f"(default: {DEFAULT_HOTEL_COLUMN})"
        ),
    )
    parser.add_argument(
        "--stay",
        action="append",
        default=None,
        metavar="DATE:ROOMS",
        help=(
            "Add a stay config entry (e.g. 2026-08-30:10). "
            f"Repeat for multiple dates. Defaults to {DEFAULT_STAY_ARGS or 'none'}."
        ),
    )
    parser.add_argument("--mcode", default=DEFAULT_MCODE, help="MCode for the booking request")
    parser.add_argument("--event-id", default=DEFAULT_EVENT_ID, help="Event ID for the booking request")
    parser.add_argument(
        "--language",
        default=DEFAULT_LANGUAGE,
        help=f"Language code for results (default: {DEFAULT_LANGUAGE})",
    )
    parser.add_argument(
        "--all-rates",
        dest="best_only",
        action="store_false",
        help="Set best_only to false (default is true)",
    )
    parser.set_defaults(best_only=True)
    parser.add_argument(
        "--url",
        default=DEFAULT_URL,
        help=f"Override bookings search URL (default: {DEFAULT_URL})",
    )
    parser.add_argument(
        "--token",
        default=os.getenv("KUONI_BOOKINGS_TOKEN") or os.getenv("KUONI_BEARER_TOKEN"),
        help="Bearer token for Authorization header (or set KUONI_BOOKINGS_TOKEN)",
    )
    parser.add_argument(
        "--auth-url",
        default=DEFAULT_AUTH_URL,
        help=f"Auth token URL for client credentials flow (default: {DEFAULT_AUTH_URL})",
    )
    parser.add_argument(
        "--auth-client-id",
        default=DEFAULT_AUTH_CLIENT_ID,
        help="Client ID for auth token fetch (default: env KUONI_AUTH_CLIENT_ID)",
    )
    parser.add_argument(
        "--auth-client-secret",
        default=DEFAULT_AUTH_CLIENT_SECRET,
        help="Client secret for auth token fetch (default: env KUONI_AUTH_CLIENT_SECRET)",
    )
    parser.add_argument(
        "--auth-audience",
        default=DEFAULT_AUTH_AUDIENCE,
        help=f"Audience for auth token fetch (default: {DEFAULT_AUTH_AUDIENCE})",
    )
    parser.add_argument(
        "--auth-grant-type",
        default=DEFAULT_AUTH_GRANT_TYPE,
        help=f"Grant type for auth token fetch (default: {DEFAULT_AUTH_GRANT_TYPE})",
    )
    parser.add_argument(
        "--jsonl-output",
        "--output",
        dest="jsonl_output",
        type=Path,
        default=DEFAULT_JSONL,
        help=f"Path to write newline-delimited JSON results (default: {DEFAULT_JSONL})",
    )
    parser.add_argument(
        "--jsonl-input",
        type=Path,
        default=None,
        help="Existing JSONL file to read when using --convert-only (defaults to --jsonl-output)",
    )
    parser.add_argument(
        "--convert-only",
        action="store_true",
        help="Skip API calls and only render outputs from the JSONL input file.",
    )
    parser.add_argument(
        "--csv-output",
        type=Path,
        default=DEFAULT_CSV,
        help=(
            "Write a CSV with hotel rows and date columns of room counts "
            f"(default: {DEFAULT_CSV})"
        ),
    )
    parser.add_argument(
        "--no-csv",
        action="store_true",
        help="Skip writing the availability CSV",
    )
    parser.add_argument(
        "--excel-output",
        default=DEFAULT_EXCEL,
        type=Path,
        help=f"Optional Excel output path (default: {DEFAULT_EXCEL})",
    )
    parser.add_argument(
        "--no-excel",
        action="store_true",
        help="Skip writing the Excel file",
    )
    parser.add_argument(
        "--png-output",
        default=DEFAULT_PNG,
        type=Path,
        help=f"Optional PNG heatmap path (default: {DEFAULT_PNG})",
    )
    parser.add_argument(
        "--png-size",
        nargs=2,
        type=int,
        metavar=("WIDTH", "HEIGHT"),
        default=DEFAULT_PNG_SIZE,
        help="PNG dimensions in pixels (default: 520 700)",
    )
    parser.add_argument(
        "--no-png",
        action="store_true",
        help="Skip writing the PNG heatmap",
    )
    parser.add_argument(
        "--timeout",
        type=float,
        default=30.0,
        help="HTTP timeout in seconds (default: 30)",
    )

    args = parser.parse_args(argv)

    hotel_map: dict[str, dict[str, int]] = {}
    jsonl_input = args.jsonl_input or args.jsonl_output
    session = requests.Session()

    # Late-bind auth defaults so .env values (loaded above) are honored even though
    # module-level constants may have been initialised before load_dotenv().
    args.auth_client_id = args.auth_client_id or os.getenv("KUONI_AUTH_CLIENT_ID")
    args.auth_client_secret = args.auth_client_secret or os.getenv("KUONI_AUTH_CLIENT_SECRET")
    args.auth_audience = args.auth_audience or os.getenv("KUONI_AUTH_AUDIENCE", DEFAULT_AUTH_AUDIENCE)
    args.auth_grant_type = args.auth_grant_type or os.getenv("KUONI_AUTH_GRANT_TYPE", DEFAULT_AUTH_GRANT_TYPE)
    args.auth_url = args.auth_url or os.getenv("KUONI_AUTH_URL", DEFAULT_AUTH_URL)

    if args.convert_only:
        if not jsonl_input.exists():
            print(f"[error] JSONL input {jsonl_input} not found", file=sys.stderr)
            return 1
        hotel_map = build_hotel_map_from_jsonl(jsonl_input)
    else:
        token = args.token
        if not token:
            # Try client credentials flow
            if args.auth_client_id and args.auth_client_secret and args.auth_url:
                try:
                    token = fetch_auth_token(
                        session,
                        url=args.auth_url,
                        client_id=args.auth_client_id,
                        client_secret=args.auth_client_secret,
                        audience=args.auth_audience,
                        grant_type=args.auth_grant_type,
                        timeout=args.timeout,
                    )
                    print("Fetched bearer token via client credentials.")
                except Exception as exc:
                    print(f"[error] Failed to fetch bearer token: {exc}", file=sys.stderr)
                    return 1
            else:
                parser.error(
                    "No bearer token provided; set --token or provide auth client credentials."
                )
        args.token = token

        stay_args = args.stay if args.stay is not None else DEFAULT_STAY_ARGS
        try:
            config = parse_stays(stay_args) if stay_args else None
            hotel_ids = read_hotel_ids(args.csv_path, args.hotel_column)
        except Exception as exc:
            print(f"[error] {exc}", file=sys.stderr)
            return 1

        hotel_map = fetch_bookings_availability(
            hotel_ids=hotel_ids,
            config=config,
            args=args,
            session=session,
        )
        jsonl_input = args.jsonl_output

    generate_outputs(
        hotel_map,
        csv_path=args.csv_output,
        excel_path=args.excel_output,
        png_path=args.png_output,
        png_size=tuple(args.png_size),
        no_csv=args.no_csv,
        no_excel=args.no_excel,
        no_png=args.no_png,
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
